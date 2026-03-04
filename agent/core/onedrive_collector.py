"""
Apollo Agent - OneDrive/SharePoint Collector (Sprint 35)
=========================================================

Cloud file collector via Microsoft Graph API.
Returns CollectorResult compatible with Agent V1.6 pipeline.

Supports:
- OneDrive Personal
- OneDrive for Business
- SharePoint Document Libraries

Requires:
- Azure AD App Registration with Files.Read.All permission
- MSAL library for OAuth authentication

(c) 2025 Gilles Gabriel - gilles.gabriel@noos.fr
"""
from __future__ import annotations

import asyncio
import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any, Callable
from pathlib import PurePosixPath

# Optional imports - checked at runtime
try:
    import msal
    import aiohttp
    CLOUD_DEPS_AVAILABLE = True
except ImportError:
    CLOUD_DEPS_AVAILABLE = False
    msal = None
    aiohttp = None

# Compatible imports
try:
    from models.contracts import FileMetadata, CollectorResult
except ImportError:
    from agent.models.contracts import FileMetadata, CollectorResult

logger = logging.getLogger(__name__)


@dataclass
class CloudPermission:
    """A single permission grant on a cloud file."""
    grantee: str           # Email or display name
    role: str              # "read", "write", "owner"
    grant_type: str        # "user", "group", "link", "application"
    link_type: Optional[str] = None  # "anonymous", "organization" (for link grants)
    inherited: bool = False


@dataclass
class CloudFileMetadata(FileMetadata):
    """Extended FileMetadata for cloud files."""
    # Cloud-specific fields
    cloud_id: Optional[str] = None  # Graph API item ID
    drive_id: Optional[str] = None
    web_url: Optional[str] = None
    is_shared: bool = False
    shared_scope: Optional[str] = None  # "anonymous", "organization", "users"
    shared_with: List[Dict[str, Any]] = field(default_factory=list)  # Detailed permissions
    created_by: Optional[str] = None
    modified_by: Optional[str] = None
    download_url: Optional[str] = None  # Temporary download URL
    # Sprint 86B Niveau 1: Graph API fields already returned, now parsed
    file_hash_sha1: Optional[str] = None  # item.file.hashes.sha1Hash
    mime_type: Optional[str] = None       # item.file.mimeType
    ctime: float = 0.0                    # item.createdDateTime → Unix timestamp
    # Sprint 86B Niveau 2: Cloud enrichment fields
    etag: Optional[str] = None            # item.eTag — entity change detection
    ctag: Optional[str] = None            # item.cTag — content change detection
    malware_detected: bool = False        # item.malware is not None
    malware_description: Optional[str] = None  # item.malware.description
    deleted_state: Optional[str] = None   # item.deleted.state
    retention_label: Optional[str] = None  # item.retentionLabel.name (DLP/MIP)


@dataclass
class CloudCollectorResult:
    """Result of cloud file collection."""
    source_type: str = "cloud"
    source_subtype: str = "onedrive"  # onedrive, sharepoint, gdrive, s3
    drive_id: Optional[str] = None
    tenant_id: Optional[str] = None
    root_path: str = "/"
    files: List[CloudFileMetadata] = field(default_factory=list)
    total_size: int = 0
    shared_files_count: int = 0
    error: Optional[str] = None
    errors: Optional[List[str]] = None

    def to_collector_result(self) -> CollectorResult:
        """Convert to standard CollectorResult for pipeline compatibility."""
        return CollectorResult(
            root_path=self.root_path,
            files=self.files,
            total_size=self.total_size,
            error=self.error,
            errors=self.errors
        )


class OneDriveCollector:
    """
    Collecteur OneDrive/SharePoint via Microsoft Graph API.
    Retourne CollectorResult compatible avec pipeline Agent V1.6.
    """
    GRAPH_API = "https://graph.microsoft.com/v1.0"
    SCOPES = ["https://graph.microsoft.com/.default"]

    # Extractible extensions (same as Hub scoring)
    EXTRACTIBLE_EXTENSIONS = {
        '.pdf', '.docx', '.xlsx', '.txt', '.csv',
        '.json', '.md', '.html', '.xml',
        '.doc', '.xls', '.pptx', '.ppt', '.odt', '.ods',
        '.rtf', '.yaml', '.yml', '.sql'
    }

    # Concurrency limit for Graph API permission calls
    PERMISSIONS_SEMAPHORE_LIMIT = 5

    def __init__(
        self,
        tenant_id: str,
        client_id: str,
        client_secret: str,
        authority: Optional[str] = None,
        fetch_permissions: bool = False
    ):
        """
        Initialize OneDrive collector with Azure AD credentials.

        Args:
            tenant_id: Azure AD Tenant ID
            client_id: Azure AD Application (Client) ID
            client_secret: Client secret value
            authority: Optional authority URL (defaults to Azure AD)
        """
        if not CLOUD_DEPS_AVAILABLE:
            raise ImportError(
                "Cloud dependencies not installed. "
                "Run: pip install msal aiohttp"
            )

        self.tenant_id = tenant_id
        self.client_id = client_id
        # client_secret passed to MSAL only, never stored as instance attribute
        self.authority = authority or f"https://login.microsoftonline.com/{tenant_id}"

        self.fetch_permissions = fetch_permissions
        self._access_token: Optional[str] = None
        self._token_expires: Optional[datetime] = None
        self._session: Optional[aiohttp.ClientSession] = None
        self._perm_semaphore: Optional[asyncio.Semaphore] = None

        # MSAL confidential client (sole holder of credential)
        self._msal_app = msal.ConfidentialClientApplication(
            client_id=client_id,
            client_credential=client_secret,
            authority=self.authority
        )

    async def authenticate(self) -> bool:
        """
        Authenticate with Microsoft Graph API using client credentials flow.

        Returns:
            True if authentication successful
        """
        try:
            result = self._msal_app.acquire_token_for_client(scopes=self.SCOPES)

            if "access_token" in result:
                self._access_token = result["access_token"]
                # Token typically valid for 1 hour
                expires_in = result.get("expires_in", 3600)
                self._token_expires = datetime.now(timezone.utc).timestamp() + expires_in
                logger.info("OneDrive authentication successful")
                return True
            else:
                error = result.get("error_description", result.get("error", "Unknown error"))
                logger.error(f"Authentication failed: {error}")
                return False

        except Exception as e:
            logger.error(f"Authentication exception: {e}")
            return False

    async def _ensure_session(self) -> aiohttp.ClientSession:
        """Get or create aiohttp session with auth headers."""
        if self._session is None or self._session.closed:
            headers = {
                "Authorization": f"Bearer {self._access_token}",
                "Content-Type": "application/json"
            }
            self._session = aiohttp.ClientSession(headers=headers)
        return self._session

    def _validate_graph_response(self, data: Any, endpoint: str) -> Dict[str, Any]:
        """Validate Graph API response is a dict and not an error payload."""
        if not isinstance(data, dict):
            raise ValueError(f"Graph API returned non-dict for {endpoint}: {type(data)}")
        if "error" in data:
            err = data["error"]
            code = err.get("code", "unknown") if isinstance(err, dict) else str(err)
            msg = err.get("message", "") if isinstance(err, dict) else ""
            raise Exception(f"Graph API error in 200 response ({code}): {msg}")
        return data

    async def _graph_request(
        self,
        endpoint: str,
        method: str = "GET",
        params: Optional[Dict] = None
    ) -> Dict[str, Any]:
        """Make authenticated request to Graph API."""
        session = await self._ensure_session()
        url = f"{self.GRAPH_API}{endpoint}"

        async with session.request(method, url, params=params) as response:
            if response.status == 200:
                data = await response.json()
                return self._validate_graph_response(data, endpoint)
            elif response.status == 401:
                # Token expired, try to refresh
                if await self.authenticate():
                    session = await self._ensure_session()
                    async with session.request(method, url, params=params) as retry:
                        if retry.status == 200:
                            data = await retry.json()
                            return self._validate_graph_response(data, endpoint)
                raise Exception(f"Authentication failed after retry")
            else:
                text = await response.text()
                raise Exception(f"Graph API error {response.status}: {text}")

    async def list_drives(self) -> List[Dict[str, Any]]:
        """
        List available SharePoint sites (client_credentials mode).

        Note: /me/drives doesn't work with client_credentials (no user context).
        Instead, we list SharePoint sites accessible to the application.

        Returns:
            List of site info dicts with id, name, webUrl
        """
        drives = []

        try:
            # List SharePoint sites (works with client_credentials)
            result = await self._graph_request("/sites?search=*")
            sites = result.get("value", [])

            for site in sites:
                site_id = site.get("id")
                if site_id:
                    # Get drives for each site
                    try:
                        site_drives = await self._graph_request(f"/sites/{site_id}/drives")
                        for d in site_drives.get("value", []):
                            drives.append({
                                "id": d.get("id"),
                                "name": f"{site.get('displayName', 'Site')} - {d.get('name', 'Drive')}",
                                "driveType": d.get("driveType", "documentLibrary"),
                                "siteId": site_id,
                                "webUrl": d.get("webUrl")
                            })
                    except Exception as e:
                        logger.warning(f"Failed to get drives for site {site_id}: {e}")

            logger.info(f"Found {len(drives)} SharePoint drives")
            return drives

        except Exception as e:
            logger.error(f"Failed to list sites/drives: {e}")
            # Return empty list with info message
            return []

    async def collect_files(
        self,
        drive_id: str = "me",
        folder_path: str = "/",
        max_files: int = 10000000,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> CloudCollectorResult:
        """
        Collect file metadata from OneDrive/SharePoint.

        Args:
            drive_id: Drive ID or "me" for user's default drive
            folder_path: Folder path to scan (default: root)
            max_files: Maximum files to collect
            progress_callback: Optional callback(count, filename) for progress

        Returns:
            CloudCollectorResult with file metadata
        """
        files: List[CloudFileMetadata] = []
        errors: List[str] = []
        total_size = 0
        shared_count = 0

        try:
            # Build endpoint based on drive_id
            if drive_id == "me":
                base_endpoint = "/me/drive"
            else:
                base_endpoint = f"/drives/{drive_id}"

            # Get root folder or specific path
            if folder_path == "/" or folder_path == "":
                endpoint = f"{base_endpoint}/root/children"
            else:
                # Encode path for URL
                clean_path = folder_path.strip("/")
                endpoint = f"{base_endpoint}/root:/{clean_path}:/children"

            # Paginated collection
            next_link = None
            file_count = 0

            while True:
                if next_link:
                    # Use full URL for pagination
                    result = await self._graph_request_full_url(next_link)
                else:
                    params = {
                        "$select": "id,name,size,file,folder,createdDateTime,lastModifiedDateTime,"
                                   "createdBy,lastModifiedBy,webUrl,parentReference,shared,"
                                   "eTag,cTag,malware,deleted,retentionLabel",
                        "$top": "200"  # Batch size
                    }
                    result = await self._graph_request(endpoint, params=params)

                items = result.get("value", [])

                for item in items:
                    if file_count >= max_files:
                        break

                    # Process item
                    if "folder" in item:
                        # Recursively scan subfolders
                        subfolder_path = f"{folder_path.rstrip('/')}/{item.get('name', 'unknown')}"
                        sub_result = await self.collect_files(
                            drive_id=drive_id,
                            folder_path=subfolder_path,
                            max_files=max_files - file_count,
                            progress_callback=progress_callback
                        )
                        files.extend(sub_result.files)
                        total_size += sub_result.total_size
                        shared_count += sub_result.shared_files_count
                        file_count += len(sub_result.files)
                        if sub_result.errors:
                            errors.extend(sub_result.errors)

                    elif "file" in item:
                        # It's a file
                        file_meta = self._parse_file_item(item, folder_path, drive_id)
                        files.append(file_meta)
                        total_size += file_meta.size
                        if file_meta.is_shared:
                            shared_count += 1
                        file_count += 1

                        if progress_callback and file_count % 100 == 0:
                            progress_callback(file_count, file_meta.name)

                # Check for more pages
                next_link = result.get("@odata.nextLink")
                if not next_link or file_count >= max_files:
                    break

            # Enrich shared files with detailed permissions if requested
            if self.fetch_permissions and shared_count > 0:
                await self._enrich_permissions(files)

            return CloudCollectorResult(
                source_type="cloud",
                source_subtype="onedrive",
                drive_id=drive_id if drive_id != "me" else None,
                tenant_id=self.tenant_id,
                root_path=folder_path,
                files=files,
                total_size=total_size,
                shared_files_count=shared_count,
                errors=errors if errors else None
            )

        except Exception as e:
            logger.error(f"Collection failed: {e}")
            return CloudCollectorResult(
                source_type="cloud",
                source_subtype="onedrive",
                drive_id=drive_id if drive_id != "me" else None,
                tenant_id=self.tenant_id,
                root_path=folder_path,
                files=files,
                total_size=total_size,
                error=str(e),
                errors=errors if errors else None
            )

    async def _graph_request_full_url(self, url: str) -> Dict[str, Any]:
        """Make request to full URL (for pagination)."""
        session = await self._ensure_session()
        async with session.get(url) as response:
            if response.status == 200:
                data = await response.json()
                return self._validate_graph_response(data, url)
            else:
                text = await response.text()
                raise Exception(f"Graph API error {response.status}: {text}")

    def _parse_file_item(
        self,
        item: Dict[str, Any],
        parent_path: str,
        drive_id: str
    ) -> CloudFileMetadata:
        """Parse Graph API item to CloudFileMetadata."""
        name = item.get("name", "unknown")
        size = item.get("size", 0)

        # Parse timestamps
        created_str = item.get("createdDateTime", "")
        modified_str = item.get("lastModifiedDateTime", "")

        try:
            mtime = datetime.fromisoformat(modified_str.replace("Z", "+00:00")).timestamp()
        except (ValueError, AttributeError):
            mtime = 0.0

        # Sprint 86B: Parse createdDateTime as ctime
        try:
            ctime = datetime.fromisoformat(created_str.replace("Z", "+00:00")).timestamp()
        except (ValueError, AttributeError):
            ctime = 0.0

        # Sprint 86B: Parse file.hashes and file.mimeType (already in $select via 'file')
        file_obj = item.get("file") or {}
        hashes = file_obj.get("hashes") or {}
        file_hash_sha1 = hashes.get("sha1Hash")
        mime_type = file_obj.get("mimeType")

        # Sprint 86B Niveau 2: Cloud enrichment fields
        etag = item.get("eTag")
        ctag = item.get("cTag")
        malware_obj = item.get("malware")
        malware_detected = malware_obj is not None
        malware_description = (malware_obj or {}).get("description")
        deleted_obj = item.get("deleted")
        deleted_state = (deleted_obj or {}).get("state")
        retention_obj = item.get("retentionLabel")
        retention_label = (retention_obj or {}).get("name") if retention_obj else None

        # Build relative path
        relative_path = f"{parent_path.strip('/')}/{name}".lstrip("/")

        # Extension
        ext = PurePosixPath(name).suffix.lower()

        # Sharing info
        shared_info = item.get("shared", {})
        is_shared = bool(shared_info)
        shared_scope = shared_info.get("scope") if is_shared else None

        # Creator/modifier (defensive: nested dicts may be None)
        created_by_obj = item.get("createdBy") or {}
        created_by = (created_by_obj.get("user") or {}).get("displayName")
        modified_by_obj = item.get("lastModifiedBy") or {}
        modified_by = (modified_by_obj.get("user") or {}).get("displayName")

        # Calculate depth from path
        depth = len(relative_path.split("/")) - 1

        return CloudFileMetadata(
            path=f"onedrive://{drive_id}/{relative_path}",
            relative_path=relative_path,
            name=name,
            extension=ext,
            size=size,
            mtime=mtime,
            depth=depth,
            # Cloud-specific
            cloud_id=item.get("id"),
            drive_id=drive_id,
            web_url=item.get("webUrl"),
            is_shared=is_shared,
            shared_scope=shared_scope,
            created_by=created_by,
            modified_by=modified_by,
            # Sprint 86B Niveau 1: Graph API fields now parsed
            file_hash_sha1=file_hash_sha1,
            mime_type=mime_type,
            ctime=ctime,
            # Sprint 86B Niveau 2: Cloud enrichment
            etag=etag,
            ctag=ctag,
            malware_detected=malware_detected,
            malware_description=malware_description,
            deleted_state=deleted_state,
            retention_label=retention_label
        )

    async def _fetch_item_permissions(
        self,
        drive_id: str,
        item_id: str
    ) -> List[Dict[str, Any]]:
        """
        Fetch detailed permissions for a single item via Graph API.

        GET /drives/{drive_id}/items/{item_id}/permissions

        Returns list of serialized CloudPermission dicts.
        """
        if not self._perm_semaphore:
            self._perm_semaphore = asyncio.Semaphore(self.PERMISSIONS_SEMAPHORE_LIMIT)

        async with self._perm_semaphore:
            try:
                if drive_id == "me":
                    endpoint = f"/me/drive/items/{item_id}/permissions"
                else:
                    endpoint = f"/drives/{drive_id}/items/{item_id}/permissions"

                result = await self._graph_request(endpoint)
                permissions = []

                for perm in result.get("value", []):
                    roles = perm.get("roles", [])
                    role = roles[0] if roles else "read"
                    inherited = perm.get("inheritedFrom") is not None

                    # User/group grant
                    granted_to = perm.get("grantedToV2") or perm.get("grantedTo") or {}
                    user = granted_to.get("user") or {}
                    group = granted_to.get("group") or {}

                    if user:
                        permissions.append({
                            "grantee": user.get("email") or user.get("displayName", "unknown"),
                            "role": role,
                            "grant_type": "user",
                            "inherited": inherited
                        })
                    elif group:
                        permissions.append({
                            "grantee": group.get("displayName", "unknown_group"),
                            "role": role,
                            "grant_type": "group",
                            "inherited": inherited
                        })

                    # Link grant
                    link = perm.get("link")
                    if link:
                        link_type = link.get("scope", "unknown")
                        permissions.append({
                            "grantee": f"link:{link_type}",
                            "role": role,
                            "grant_type": "link",
                            "link_type": link_type,
                            "inherited": inherited
                        })

                    # Application grant
                    app = (granted_to.get("application") or
                           perm.get("grantedToIdentitiesV2", [{}])[0].get("application") if
                           perm.get("grantedToIdentitiesV2") else None)
                    if app and not user and not group:
                        permissions.append({
                            "grantee": app.get("displayName", "unknown_app"),
                            "role": role,
                            "grant_type": "application",
                            "inherited": inherited
                        })

                return permissions

            except Exception as e:
                logger.warning(f"Failed to fetch permissions for item {item_id}: {e}")
                return []

    async def _enrich_permissions(
        self,
        files: List[CloudFileMetadata]
    ) -> int:
        """
        Enrich shared files with detailed permission data.

        Only fetches permissions for files where is_shared=True.
        Uses semaphore to limit concurrent Graph API calls.

        Returns count of files enriched.
        """
        shared_files = [f for f in files if f.is_shared and f.cloud_id]
        if not shared_files:
            return 0

        logger.info(f"Fetching permissions for {len(shared_files)} shared files...")

        async def enrich_one(file_meta: CloudFileMetadata):
            perms = await self._fetch_item_permissions(
                drive_id=file_meta.drive_id or "me",
                item_id=file_meta.cloud_id
            )
            file_meta.shared_with = perms

        tasks = [enrich_one(f) for f in shared_files]
        await asyncio.gather(*tasks)

        enriched = sum(1 for f in shared_files if f.shared_with)
        logger.info(f"Permissions enriched: {enriched}/{len(shared_files)} files")
        return enriched

    async def download_file_content(
        self,
        drive_id: str,
        file_id: str,
        max_bytes: int = 4096
    ) -> Optional[bytes]:
        """
        Download first N bytes of file content for PII scanning.

        Uses @microsoft.graph.downloadUrl (2-step) instead of /content (302
        redirect) to avoid intermittent failures from Authorization header
        being forwarded to Azure Blob Storage.

        Args:
            drive_id: Drive ID
            file_id: File ID from cloud_id
            max_bytes: Maximum bytes to download

        Returns:
            File content bytes or None on error
        """
        try:
            # Step 1: Get pre-authenticated download URL via Graph API
            if drive_id == "me":
                endpoint = f"/me/drive/items/{file_id}?select=id,@microsoft.graph.downloadUrl"
            else:
                endpoint = f"/drives/{drive_id}/items/{file_id}?select=id,@microsoft.graph.downloadUrl"

            session = await self._ensure_session()
            url = f"{self.GRAPH_API}{endpoint}"

            async with session.get(url) as response:
                if response.status != 200:
                    logger.warning(f"[DIAG-DL] step1 FAIL status={response.status} file_id={file_id[:12]}")
                    return None
                data = await response.json()

            download_url = data.get("@microsoft.graph.downloadUrl")
            if not download_url:
                logger.warning(f"[DIAG-DL] step1 NO downloadUrl, keys={list(data.keys())} file_id={file_id[:12]}")
                return None
            logger.warning(f"[DIAG-DL] step1 OK file_id={file_id[:12]} url_len={len(download_url)}")

            # Step 2: Download from pre-auth URL WITHOUT Authorization header
            async with aiohttp.ClientSession() as clean_session:
                async with clean_session.get(download_url) as dl_resp:
                    if dl_resp.status == 200:
                        content = await dl_resp.read()  # Full body read, not partial buffer
                        logger.warning(f"[DIAG-DL] step2 OK file_id={file_id[:12]} bytes={len(content)}")
                        return content[:max_bytes] if len(content) > max_bytes else content
                    else:
                        logger.warning(f"[DIAG-DL] step2 FAIL status={dl_resp.status} file_id={file_id[:12]}")
                        return None

        except Exception as e:
            logger.error(f"Download error: {e}")
            return None

    async def close(self):
        """Close HTTP session."""
        if self._session and not self._session.closed:
            await self._session.close()

    def __del__(self):
        """Cleanup on deletion."""
        if self._session and not self._session.closed:
            try:
                asyncio.get_event_loop().run_until_complete(self._session.close())
            except Exception:
                pass


# =============================================================================
# TEXT EXTRACTION FROM CLOUD FILE BYTES (Sprint 85)
# =============================================================================

# Extensions extractible as plain text (UTF-8 decode)
TEXT_EXTENSIONS = {
    '.csv', '.txt', '.json', '.xml', '.html', '.htm',
    '.md', '.rst', '.log', '.sql', '.yaml', '.yml',
    '.py', '.js', '.ts', '.ini', '.cfg', '.conf', '.env',
    '.properties', '.rtf'
}

# Max bytes to download for PII scan (5MB)
MAX_DOWNLOAD_BYTES = 5 * 1024 * 1024


def extract_text_from_bytes(content: bytes, extension: str) -> Optional[str]:
    """
    Extract readable text from file bytes based on extension.

    Sprint 85: Supports text files (UTF-8 decode) and XLSX (openpyxl).

    Args:
        content: Raw file bytes downloaded from Graph API
        extension: File extension (e.g. '.csv', '.xlsx')

    Returns:
        Extracted text content, or None if format not supported
    """
    ext = extension.lower()

    # Text-based formats: direct decode
    if ext in TEXT_EXTENSIONS:
        for encoding in ('utf-8', 'latin-1', 'cp1252'):
            try:
                return content.decode(encoding)
            except (UnicodeDecodeError, ValueError):
                continue
        return content.decode('utf-8', errors='ignore')

    # XLSX: extract cell values via openpyxl
    if ext in ('.xlsx', '.xls'):
        return _extract_xlsx_text(content)

    return None


def _extract_xlsx_text(content: bytes) -> Optional[str]:
    """
    Extract text from XLSX bytes using openpyxl.

    Reads all sheets, all rows, concatenates cell values as text lines.
    """
    try:
        from io import BytesIO
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        text_parts = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = ' '.join(
                    str(cell) for cell in row if cell is not None
                )
                if row_text.strip():
                    text_parts.append(row_text)

        wb.close()
        return '\n'.join(text_parts) if text_parts else None

    except ImportError:
        logger.warning("openpyxl not installed, skipping XLSX extraction")
        return None
    except Exception as e:
        logger.warning(f"XLSX extraction failed: {e}")
        return None


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def check_cloud_dependencies() -> bool:
    """Check if cloud dependencies are installed."""
    return CLOUD_DEPS_AVAILABLE


async def collect_onedrive_files(
    tenant_id: str,
    client_id: str,
    client_secret: str,
    drive_id: str = "me",
    folder_path: str = "/",
    max_files: int = 10000000,
    progress_callback: Optional[Callable[[int, str], None]] = None,
    fetch_permissions: bool = False
) -> CloudCollectorResult:
    """
    High-level function to collect OneDrive files.

    Convenience wrapper around OneDriveCollector.

    Args:
        fetch_permissions: If True, fetch detailed permissions (grantees, roles)
                          for shared files via additional Graph API calls.
    """
    collector = OneDriveCollector(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        fetch_permissions=fetch_permissions
    )

    try:
        if not await collector.authenticate():
            return CloudCollectorResult(
                source_type="cloud",
                source_subtype="onedrive",
                error="Authentication failed"
            )

        result = await collector.collect_files(
            drive_id=drive_id,
            folder_path=folder_path,
            max_files=max_files,
            progress_callback=progress_callback
        )

        return result

    finally:
        await collector.close()
